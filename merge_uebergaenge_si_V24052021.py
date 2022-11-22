#! /usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import re
import os
import subprocess
import glob
import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl as op
from pandas import ExcelWriter
from pandas import ExcelFile
from tempfile import TemporaryFile


# -------------------------------------------------------------------------- #
# ---------------------- Interaktion mit dem Benutzer ---------------------- #
# Der Anwender muss einen Ordner erstellen, in dem diese .py-Datei und die
# Schlüsseltabelle (Zuordnung der Schulen zu Bildungsregionen) ist. Dieser
# Ordner enthält daneben einen Unterorder, in dem die zu importierenden Excel-
# Dateien liegen. In diesem Unterordner dürfen sich keine anderen Dateien
# befinden! Stellen Sie sicher, dass keine der Dateien geöffnet ist!

"""
print("Willkommen!\n")

print("Der Anwender muss einen Ordner erstellen, in dem diese .py-Datei "
      "und die Schlüsseltabelle (Zuordnung der Schulen zu Bildungsregionen) "
      "ist. Dieser Ordner enthält daneben einen Unterorder, in dem die zu "
      "importierenden Excel-Dateien liegen. In diesem Unterordner dürfen "
      "sich keine anderen Dateien befinden! \n"
      "Stellen Sie sicher, dass keine der Dateien geöffnet ist und dass "
      "alle die Endung .xlsx haben!\n")

xfolder = input(
    "Wie heißt der Ordner in dem die zu importierenden Excel-Dateien liegen? "
    "Wenn der Ordner > 0_SuS_Uebergang_SekI < heißt, drücken Sie die "
    "Eingabetaste!\n") or "0_SuS_Uebergang_SekI"
regfile = input(
    "Wie heißt die Excel-Datei [mit Endung!], in der die Schlüsseltabelle "
    "hinterlegt ist? Wenn es sich um die Tabelle > SCHL_Schulen.xlsx < handelt, "
    "drücken Sie die Eingabetaste!\n") or "SCHL_Schulen.xlsx"
cols_region = input(
    "Wie heißt die Spalte mit der Schulnummer und der Bildungsregion? "
    "Bitte geben Sie diese mit Leerzeichen getrennt ein [z.B. Schulnr. BR]. "
    "Wenn die Spalten > Schulnr. BR < heißen, drücken Sie die "
    "Eingabetaste!\n") or "Schulnr. BR"

print("Bitte warten Sie, das Programm wird ausgeführt! Das kann einige Zeit "
      "in Anspruch nehmen!\n")

"""
xfolder = "0_SuS_Uebergang_SekI"
regfile = "SCHL_Schulen.xlsx"
cols_region = "Schulnr. BR"


# -------------------------------------------------------------------------- #
# -------------------------------------------------------------------------- #
# Liste der Dateien in dem Ordner mit den zu importierenden Dateien

# Aktueller Pfad
xtype = "xlsx"
cwd = os.getcwd()
xfiles = sorted(glob.glob(cwd+"/"+xfolder+"/*."+xtype))
#os.chdir()

# -------------------------------------------------------------------------- #
# Informationen aus der Schlüsseltabelle - Bildungsregionen und Schulen
# headsheets = []
column_id = cols_region.split(" ")[0]
column_reg = cols_region.split(" ")[1]

reg = pd.read_excel(regfile, engine="openpyxl")
reg = reg.dropna(axis=1, how='all')
headerstr = list(reg.columns.values)

idxschnr = headerstr.index(column_id)
idxreg = headerstr.index(column_reg)

idnr_pre = list(reg[headerstr[idxschnr]])
regions_pre = list(reg[headerstr[idxreg]])
    
regid = []
regname = []

for i, re0 in enumerate(regions_pre):
    if pd.isna(re0) != True:
        regname.append(re0)
        regid.append(idnr_pre[i])
    if pd.isna(re0) != False and idnr_pre[i] < 79999:
        print("Schulnummer "+str(idnr_pre[i])+" wurde in der Schlüsseltabele "
              +regfile+" keine Bildungsregion zugeordnet! Bitte erledigen "
              "Sie das zuerst! Führen Sie das Programm danach erneut aus!\n")
        sys.exit(1)

# -------------------------------------------------------------------------- #
# Finale Tabelle formatieren
writer = pd.ExcelWriter("Uebergaenge_SekI_ganzFFM.xlsx", engine='xlsxwriter')
workbook = writer.book
sheetname = "uebergang_seki_raw"
worksheet = workbook.add_worksheet(sheetname)
writer.sheets[sheetname] = worksheet
worksheet.set_column('A:AK', 12)
header_format = workbook.add_format(
    {'bold': 1, 'border': 0, 'align': 'center', 'valign': 'vcenter',
     'fg_color': '#FFFFFF', 'bg_color': 'solid'})

# -------------------------------------------------------------------------- #
# Eingangstabellen überprüfen
# Eigentliche Überschrift/Kopfzeile finden, dort beginnt die Tabelle
# Kopfzeile überspringen und uniqueness sicherstellen
# Schuljahr identifizieren und Indizieren
years = {} 
indices = {}
headers = {}
secondheaders = {}
sheets = {}

# Die Sheets mit den Daten müssen alle mit > 5-F < beginnen
for xf in xfiles:
    nametemp = xf.split("/")[-1].split(".")[0]+".temp"
    namefile = xf.split("/")[-1]
    # Diese Methode ist besser für Dateien mit mehreren Sheets
    wb = op.load_workbook(xf)
    #indf = pd.ExcelFile(xf)
    # Generiere ein Sheet-Dictionary
    all_sheets_pre = wb.sheetnames
    all_sheets = []
    for s in all_sheets_pre:
        if "5-" in s:
            all_sheets.append(s)
    sheets[namefile] = all_sheets

    headsheets = []
    for s in all_sheets:
        wi = wb.worksheets.index(wb[s])
        ws = wb[wb.sheetnames[wi]]
        df = pd.DataFrame(ws.values)
        #df = pd.read_excel(xf,sheet_name=s)
        
        hdf_find = df[df[list(df)[0]].str.contains("Jg.",na=False)]
        hdf_pre = hdf_find[list(hdf_find)[0:]]
        hdf = hdf_pre.values.tolist()[0]
        headsheets.append(hdf)
    headers[namefile] = headsheets
    
    wb = op.load_workbook(xf)
    wi = wb.worksheets.index(wb[all_sheets[0]])
    ws = wb[wb.sheetnames[wi]]
    df = pd.DataFrame(ws.values)
    #df = pd.read_excel(xf,sheet_name=all_sheets[0], engine="openpyxl")
    df.to_csv(nametemp,sep=';')

    m = 0
    with open(nametemp, 'r') as csv:
        for l in csv:
            line = l.rstrip().split(";")
            cases = ["Name" in line, "PLZ" in line]
            if all(cases):
                #headers[namefile] = line[1:]
                indices[namefile] = int(line[0])+4
                m = 1
            if m == 1 and not any(cases) :
                #secondheaders[namefile] = line[1:]
                m = 0
            if m == 0:
                if "Parameterliste" in l:
                    for i, s in enumerate(line):
                        if "Schuljahr" in s:						
                            # Es muss eine Zeile existieren, die mit dem 
							# Wort "Parameterliste" beginnt
							# Beispiel:
                            # Parameterliste: Schuljahr: 2018/2019; 
							# Schulaufsicht: F; Stufe: 5, 7, 11
                            year = line[i].split(":")[-1].split("/")[0]
                            years[namefile] = year
 

# Eigentliche Überschrift/Kopfzeile finden, dort beginnt die Tabelle
newheaders = []
for xf in xfiles:
    namefile = xf.split("/")[-1]
    for h in headers[namefile]:
        newh = [x.rstrip() for x in h if str(x) != 'nan' and str(x) != 'None']
        newheaders.append(newh)
lheader = newheaders[0]
inds = 0
for h in newheaders[1:]:
    for i,ele in enumerate(h):
        ele = str(ele).rstrip()
        if ele == "S":
            inds = i
        if ele not in lheader:
            lheader.append(ele)
for j in range(0, len(lheader)-inds):
    real = str(lheader[inds+(j*2)])
    lheader[inds+(j*2)] = str(lheader[inds+(j*2)])+"_abs"
    lheader.insert(inds+(j*2)+1, real+"_pro")

idxnam1 = [i for i, n in enumerate(lheader) if n == "Name"][0]
lheader.insert(idxnam1+1,"BR")
idxnam2 = [i for i, n in enumerate(lheader) if n == "Name"][1]
lheader.insert(idxnam2+1,"BR")
lheader[idxnam1] = "Name1"
lheader[idxnam2] = "Name2"

idxid1 = [i for i, n in enumerate(lheader) if n == "Nr."][0]
idxid2 = [i for i, n in enumerate(lheader) if n == "Nr."][1]
idxop1 = [i for i, n in enumerate(lheader) if n == "Ö/P"][0]
idxop2 = [i for i, n in enumerate(lheader) if n == "Ö/P"][1]
idxtyp1 = [i for i, n in enumerate(lheader) if n == "Typ"][0]
idxtyp2 = [i for i, n in enumerate(lheader) if n == "Typ"][1]
idxplz1 = [i for i, n in enumerate(lheader) if n == "PLZ"][0]
idxplz2 = [i for i, n in enumerate(lheader) if n == "PLZ"][1]
idxort1 = [i for i, n in enumerate(lheader) if n == "Ort"][0]
idxort2 = [i for i, n in enumerate(lheader) if n == "Ort"][1]
lheader[idxid1] = "Nr_1"
lheader[idxid2] = "Nr_2"
lheader[idxop1] = "Ö/P_1"
lheader[idxop2] = "Ö/P_2"
lheader[idxtyp1] = "Typ_1"
lheader[idxtyp2] = "Typ_2"
lheader[idxplz1] = "PLZ_1"
lheader[idxplz2] = "PLZ_2"
lheader[idxort1] = "Ort_1"
lheader[idxort2] = "Ort_2"

lheader_orig = []
for ele in lheader:
    if ele != "BR":
        lheader_orig.append(ele)
#print(lheader_orig)

lheader.insert(0, "Jahr")


# -------------------------------------------------------------------------- #
# -------------------------------------------------------------------------- #
# Programm starten, Tabellen aller Jahre untereinander hängen, BR zuordnen

# ------------------- LOOP ------------------------------------------------- #
row = 1
lheader_ini = 0
header_def = []
for j, xf in enumerate(xfiles):
    namefile = xf.split("/")[-1]
    print("In "+namefile+" wird gerade gearbeitet!\n")
    
    # LOOP über Tabellenblätter -------------------------------------------- #
    list_sheets = sheets[namefile]
	
    for k, ws in enumerate(list_sheets):            
        h = headers[namefile][k]
        # Neue Spaltenüberschriften
        hup = []
        for i, el in enumerate(h):
            if str(el) == 'nan' or str(el) == 'None':
                    hup.append("")
            else:
                    hup.append(el)
        # Index für leere Felder
        idxes = [i for i, n in enumerate(hup) if n == ""]
        for i in idxes:
            if str(h[i-1]) != 'nan' and str(h[i-1]) != 'None':
                hup[i] = str(hup[i-1]).rstrip()+"_pro"
                hup[i-1] = str(hup[i-1]).rstrip()+"_abs"    
        
        # Suche Spalten
        idxnam1 = [i for i, n in enumerate(hup) if n == "Name"][0]
        idxnam2 = [i for i, n in enumerate(hup) if n == "Name"][1]
        idxid1 = [i for i, n in enumerate(hup) if n == "Nr."][0]
        idxid2 = [i for i, n in enumerate(hup) if n == "Nr."][1]      
        idxop1 = [i for i, n in enumerate(hup) if n == "Ö/P"][0]
        idxop2 = [i for i, n in enumerate(hup) if n == "Ö/P"][1]
        idxtyp1 = [i for i, n in enumerate(hup) if n == "Typ"][0]
        idxtyp2 = [i for i, n in enumerate(hup) if n == "Typ"][1]
        idxplz1 = [i for i, n in enumerate(hup) if n == "PLZ"][0]
        idxplz2 = [i for i, n in enumerate(hup) if n == "PLZ"][1]
        idxort1 = [i for i, n in enumerate(hup) if n == "Ort"][0]
        idxort2 = [i for i, n in enumerate(hup) if n == "Ort"][1]

        # Ersetze mit eindeutigen Überschriften
        hup[idxnam1] = "Name1"
        hup[idxnam2] = "Name2"
        hup[idxid1] = "Nr_1"
        hup[idxid2] = "Nr_2"
        hup[idxop1] = "Ö/P_1"
        hup[idxop2] = "Ö/P_2"
        hup[idxtyp1] = "Typ_1"
        hup[idxtyp2] = "Typ_2"
        hup[idxplz1] = "PLZ_1"
        hup[idxplz2] = "PLZ_2"
        hup[idxort1] = "Ort_1"
        hup[idxort2] = "Ort_2"
        #print(hup) 
		
        # Überspringe die Kopfzeile
        nr_skip = int(indices[namefile])-1
        listskip = [x for x in range(0,nr_skip)]
              
        newhup = []
        nulindex = []
        for i, el in enumerate(lheader_orig):
            if el in hup:
                newhup.append(el)
            else:
                nulindex.append(i)
        #print(newhup)
        df = pd.read_excel(xf, sheet_name=ws, names=hup, header=None, index_col=None, skiprows=listskip, engine="openpyxl")
        df = df.dropna(axis=1,how='all')
        df = df.dropna(axis=0,how='all')
        df = df[newhup]
        
        rowsdf = df.shape[0]
        nulcol = ["" for x in range(0, rowsdf)]
        for ind in nulindex:
            df.insert(ind, "nul", nulcol, True)
        firstcol = [int(years[namefile]) for x in range(0, rowsdf)]
        df.insert(0, "Jahr", firstcol, True)
        
        # Identifiziere Schulnr. und ordne BR zu
        colid1 = list(df["Nr_1"])
        colid2 = list(df["Nr_2"])
        
        newreg1 = []
        for i, id1 in enumerate(colid1):
            catch1 = 0
            for j,ir in enumerate(regid):
                if id1 == ir:
                    newreg1.append(regname[j])
                    catch1 = 1
            if catch1 == 0:
                newreg1.append("NA")
        newreg2 = []
        for i,id2 in enumerate(colid2):
            catch2 = 0
            for j, ir in enumerate(regid):
                if id2 == ir:
                    newreg2.append(regname[j])
                    catch2 = 1
            if catch2 == 0:
                newreg2.append("NA")
        
        newheader = list(df.columns.values)

                
        
        #print(newheader)
        idxnam1 = [i for i, n in enumerate(newheader) if n == "Name1"][0]
        df.insert(idxnam1+1, "BR", newreg1, True)
        newheader = list(df.columns.values)
        idxnam2 = [i for i, n in enumerate(newheader) if n == "Name2"][0]
        df.insert(idxnam2+1, "BR", newreg2, True)
        newheader = list(df.columns.values)
        df_obj = df.select_dtypes(['object'])
        df[df_obj.columns] = df_obj.apply(lambda x: x.str.strip())
        df.to_excel(writer,sheet_name=sheetname,startrow=row,startcol=0,index=False,header=False,float_format = "%0.1f") 
        row = row + rowsdf  
        #print(newheader)
        ####Ich suche hier nach der neuesten Version der Header, die alle Elemente enthält
        newheader_pre = []
        for elh in newheader:
            if elh != "nul":
                newheader_pre.append(elh)
        lheader = len(newheader_pre)
        if lheader > lheader_ini:
            header_def = newheader
            lheader_ini = lheader

# Neue Spaltenüberschriften in die Tabelle schreiben
for i,h in enumerate(header_def):
    worksheet.write(0, i, h, header_format)
writer.save()

# Löscht temporäre Dateien
pretempfiles = sorted(glob.glob(xfolder+"/"+"*.temp"))
for p in pretempfiles:
    os.remove(p)
print("")
print("Die Datei wurde erstellt!")
input("Drücken Sie die Eingabetaste um das Fenster zu schließen!")
